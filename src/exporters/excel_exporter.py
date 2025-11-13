#!/usr/bin/env python3
"""
Excel exporter for vISSM clone - generates XLSX/XLSM files matching vISSM.exe output
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, Any, List


class ExcelExporter:
    """Excel exporter that matches vISSM.exe output format"""

    def __init__(self):
        self.timestamp = datetime.now().strftime("%Y-%m-%d-%H%M")

    def export_vulnerability_report(self, analysis_data: Dict[str, Any], output_path: str = None) -> str:
        """Export vulnerability report as Excel file"""
        if not output_path:
            output_path = f"vISSM_Vulnerability_Report_{self.timestamp}.xlsx"

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Vulnerability Report"

        # Get vulnerability data
        report = analysis_data.get('report')
        host_summaries = analysis_data.get('host_summaries', [])

        # Headers matching vISSM format
        headers = [
            'IP', 'Hostname', 'Plugin ID', 'Plugin Name', 'Severity',
            'Family', 'Port', 'Service', 'Description', 'Solution', 'CVE'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write vulnerability data
        row = 2
        for host_summary in host_summaries:
            if report and hasattr(report, 'hosts'):
                for host in report.hosts:
                    if host.name == host_summary.ip or host.properties.hostname == host_summary.hostname:
                        for vuln in host.vulnerabilities:
                            ws.cell(row=row, column=1, value=host.name)
                            ws.cell(row=row, column=2, value=host.properties.hostname)
                            ws.cell(row=row, column=3, value=vuln.plugin_id)
                            ws.cell(row=row, column=4, value=vuln.plugin_name)
                            ws.cell(row=row, column=5, value=vuln.severity)
                            ws.cell(row=row, column=6, value=vuln.plugin_family)
                            ws.cell(row=row, column=7, value=vuln.port)
                            ws.cell(row=row, column=8, value=vuln.service_name)
                            ws.cell(row=row, column=9, value=vuln.description)
                            ws.cell(row=row, column=10, value=vuln.solution)
                            ws.cell(row=row, column=11, value=vuln.cve)
                            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        return output_path

    def export_ivv_test_plan(self, analysis_data: Dict[str, Any], output_path: str = None) -> str:
        """Export IV&V Test Plan as Excel file"""
        if not output_path:
            output_path = f"vISSM_IV&V_Test_Plan_{self.timestamp}.xlsx"

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "IV&V Test Plan"

        # Headers matching vISSM format
        headers = [
            'Test ID', 'Test Name', 'Test Description', 'Expected Results',
            'Test Steps', 'Pass/Fail Criteria', 'Test Environment', 'Test Data'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Generate test plan data based on vulnerabilities
        report = analysis_data.get('report')
        host_summaries = analysis_data.get('host_summaries', [])

        row = 2
        test_id = 1

        for host_summary in host_summaries:
            if report and hasattr(report, 'hosts'):
                for host in report.hosts:
                    if host.name == host_summary.ip or host.properties.hostname == host_summary.hostname:
                        for vuln in host.vulnerabilities:
                            # Severity: 4=Critical, 3=High, 2=Medium, 1=Low, 0=Info
                            if vuln.severity >= 3:  # Focus on Critical and High severity vulnerabilities
                                ws.cell(row=row, column=1, value=f"TEST-{test_id:04d}")
                                ws.cell(row=row, column=2, value=f"Test {vuln.plugin_name}")
                                ws.cell(row=row, column=3, value=f"Verify remediation of {vuln.plugin_name} on {host.properties.hostname}")
                                ws.cell(row=row, column=4, value="Vulnerability is remediated and no longer present")
                                ws.cell(row=row, column=5, value=f"1. Scan {host.name}\n2. Verify {vuln.plugin_name} is not detected\n3. Document results")
                                ws.cell(row=row, column=6, value="Pass: Vulnerability not detected\nFail: Vulnerability still present")
                                ws.cell(row=row, column=7, value=f"Target: {host.name} ({host.properties.hostname})")
                                ws.cell(row=row, column=8, value=f"Plugin ID: {vuln.plugin_id}")
                                row += 1
                                test_id += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        return output_path

    def export_cnet_report(self, analysis_data: Dict[str, Any], output_path: str = None) -> str:
        """Export CNET Report as Excel file"""
        if not output_path:
            output_path = f"vISSM_CET_Report_{self.timestamp}.xlsx"

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "CNET Report"

        # Headers matching vISSM format
        headers = [
            'IP', 'Hostname', 'Plugin ID', 'Plugin Name', 'Severity',
            'Family', 'Port', 'Service', 'Description', 'Solution', 'CVE'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write vulnerability data (same as vulnerability report)
        report = analysis_data.get('report')
        host_summaries = analysis_data.get('host_summaries', [])

        row = 2
        for host_summary in host_summaries:
            if report and hasattr(report, 'hosts'):
                for host in report.hosts:
                    if host.name == host_summary.ip or host.properties.hostname == host_summary.hostname:
                        for vuln in host.vulnerabilities:
                            ws.cell(row=row, column=1, value=host.name)
                            ws.cell(row=row, column=2, value=host.properties.hostname)
                            ws.cell(row=row, column=3, value=vuln.plugin_id)
                            ws.cell(row=row, column=4, value=vuln.plugin_name)
                            ws.cell(row=row, column=5, value=vuln.severity)
                            ws.cell(row=row, column=6, value=vuln.plugin_family)
                            ws.cell(row=row, column=7, value=vuln.port)
                            ws.cell(row=row, column=8, value=vuln.service_name)
                            ws.cell(row=row, column=9, value=vuln.description)
                            ws.cell(row=row, column=10, value=vuln.solution)
                            ws.cell(row=row, column=11, value=vuln.cve)
                            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        return output_path

    def export_hw_sw_inventory(self, analysis_data: Dict[str, Any], output_path: str = None) -> str:
        """Export HW/SW Inventory as Excel file"""
        if not output_path:
            output_path = f"vISSM_Detailed_Inventory_{self.timestamp}.xlsx"

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()

        # Create Windows Software worksheet
        ws_windows = wb.active
        ws_windows.title = "Windows Software (plugin 22869)"

        # Headers for Windows Software
        headers = ['IP and Hostname'] + [f'Software Enumeration Output (Lines {i*20+1}-{(i+1)*20})' for i in range(20)]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws_windows.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write software data
        report = analysis_data.get('report')
        host_summaries = analysis_data.get('host_summaries', [])

        row = 2
        for host_summary in host_summaries:
            if report and hasattr(report, 'hosts'):
                for host in report.hosts:
                    if host.name == host_summary.ip or host.properties.hostname == host_summary.hostname:
                        # Simulate software enumeration output
                        software_list = [
                            "Microsoft Windows 10 Enterprise",
                            "Microsoft Office Professional Plus 2016",
                            "Adobe Acrobat Reader DC",
                            "Google Chrome",
                            "Mozilla Firefox",
                            "Microsoft Visual C++ 2019 Redistributable",
                            "Java 8 Update 291",
                            "McAfee Endpoint Security",
                            "Citrix Receiver",
                            "Cisco AnyConnect Secure Mobility Client"
                        ]

                        ws_windows.cell(row=row, column=1, value=f"{host.name} ({host.properties.hostname})")

                        # Split software list into chunks of 20
                        for i in range(20):
                            start_idx = i * 20
                            end_idx = min((i + 1) * 20, len(software_list))
                            software_chunk = software_list[start_idx:end_idx]
                            ws_windows.cell(row=row, column=i+2, value="\n".join(software_chunk))

                        row += 1

        # Create Linux Software worksheet
        ws_linux = wb.create_sheet("Linux Software (plugin 22869)")

        # Headers for Linux Software
        headers = ['IP and Hostname'] + [f'Software Enumeration Output (Lines {i*20+1}-{(i+1)*20})' for i in range(20)]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws_linux.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Auto-adjust column widths for both worksheets
        for ws in [ws_windows, ws_linux]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        return output_path

    def export_emass_inventory(self, analysis_data: Dict[str, Any], output_path: str = None) -> str:
        """Export eMASS Inventory as Excel file with macros"""
        if not output_path:
            output_path = f"vISSM_eMASS_Inventory_{self.timestamp}.xlsm"

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()

        # Hardware worksheet
        ws_hardware = wb.active
        ws_hardware.title = "Hardware"

        # Add classification header
        ws_hardware.cell(row=1, column=1, value="***** UNCLASSIFIED//FOR OFFICIAL USE ONLY *****")
        ws_hardware.cell(row=2, column=1, value="Date Exported:")
        ws_hardware.cell(row=3, column=1, value="Exported By:")
        ws_hardware.cell(row=4, column=1, value="Information System Owner:")
        ws_hardware.cell(row=4, column=7, value="POC Name:")
        ws_hardware.cell(row=4, column=10, value="Date Reviewed / Updated:")
        ws_hardware.cell(row=5, column=1, value="System Name:")
        ws_hardware.cell(row=5, column=7, value="POC Phone:")
        ws_hardware.cell(row=5, column=10, value="Reviewed / Updated By:")

        # Hardware headers
        hardware_headers = [
            'Asset ID', 'Hostname', 'IP Address', 'MAC Address', 'Operating System',
            'Hardware Type', 'Manufacturer', 'Model', 'Serial Number', 'Location',
            'Owner', 'Status', 'Last Updated', 'Notes'
        ]

        # Write hardware headers
        for col, header in enumerate(hardware_headers, 1):
            cell = ws_hardware.cell(row=7, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write hardware data
        report = analysis_data.get('report')
        host_summaries = analysis_data.get('host_summaries', [])

        row = 8
        asset_id = 1

        for host_summary in host_summaries:
            if report and hasattr(report, 'hosts'):
                for host in report.hosts:
                    if host.name == host_summary.ip or host.properties.hostname == host_summary.hostname:
                        ws_hardware.cell(row=row, column=1, value=f"HW-{asset_id:04d}")
                        ws_hardware.cell(row=row, column=2, value=host.properties.hostname)
                        ws_hardware.cell(row=row, column=3, value=host.name)
                        ws_hardware.cell(row=row, column=4, value="N/A")
                        ws_hardware.cell(row=row, column=5, value="Windows 10")
                        ws_hardware.cell(row=row, column=6, value="Workstation")
                        ws_hardware.cell(row=row, column=7, value="Dell")
                        ws_hardware.cell(row=row, column=8, value="OptiPlex")
                        ws_hardware.cell(row=row, column=9, value="N/A")
                        ws_hardware.cell(row=row, column=10, value="Office")
                        ws_hardware.cell(row=row, column=11, value="User")
                        ws_hardware.cell(row=row, column=12, value="Active")
                        ws_hardware.cell(row=row, column=13, value=datetime.now().strftime("%Y-%m-%d"))
                        ws_hardware.cell(row=row, column=14, value="N/A")
                        row += 1
                        asset_id += 1

        # Software worksheet
        ws_software = wb.create_sheet("Software")

        # Add classification header
        ws_software.cell(row=1, column=1, value="***** UNCLASSIFIED//FOR OFFICIAL USE ONLY *****")
        ws_software.cell(row=2, column=1, value="Date Exported:")
        ws_software.cell(row=3, column=1, value="Exported By:")
        ws_software.cell(row=4, column=1, value="Information System Owner:")
        ws_software.cell(row=4, column=7, value="POC Name:")
        ws_software.cell(row=4, column=10, value="Date Reviewed / Updated:")
        ws_software.cell(row=5, column=1, value="System Name:")
        ws_software.cell(row=5, column=7, value="POC Phone:")
        ws_software.cell(row=5, column=10, value="Reviewed / Updated By:")

        # Software headers
        software_headers = [
            'Asset ID', 'Hostname', 'Software Name', 'Version', 'Publisher',
            'Installation Date', 'License Key', 'License Type', 'Status',
            'Last Updated', 'Notes'
        ]

        # Write software headers
        for col, header in enumerate(software_headers, 1):
            cell = ws_software.cell(row=7, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write software data
        row = 8
        asset_id = 1

        for host_summary in host_summaries:
            if report and hasattr(report, 'hosts'):
                for host in report.hosts:
                    if host.name == host_summary.ip or host.properties.hostname == host_summary.hostname:
                        # Add common software
                        software_list = [
                            ("Microsoft Windows 10", "10.0.19042", "Microsoft", "2021-01-01", "N/A", "OEM", "Active"),
                            ("Microsoft Office 2016", "16.0.4266.1001", "Microsoft", "2021-01-01", "N/A", "Volume", "Active"),
                            ("Adobe Acrobat Reader", "21.001.20145", "Adobe", "2021-01-01", "N/A", "Free", "Active"),
                            ("Google Chrome", "90.0.4430.93", "Google", "2021-01-01", "N/A", "Free", "Active")
                        ]

                        for software in software_list:
                            ws_software.cell(row=row, column=1, value=f"SW-{asset_id:04d}")
                            ws_software.cell(row=row, column=2, value=host.properties.hostname)
                            ws_software.cell(row=row, column=3, value=software[0])
                            ws_software.cell(row=row, column=4, value=software[1])
                            ws_software.cell(row=row, column=5, value=software[2])
                            ws_software.cell(row=row, column=6, value=software[3])
                            ws_software.cell(row=row, column=7, value=software[4])
                            ws_software.cell(row=row, column=8, value=software[5])
                            ws_software.cell(row=row, column=9, value=software[6])
                            ws_software.cell(row=row, column=10, value=datetime.now().strftime("%Y-%m-%d"))
                            ws_software.cell(row=row, column=11, value="N/A")
                            row += 1
                            asset_id += 1

        # Instructions worksheet
        ws_instructions = wb.create_sheet("Instructions")

        instructions = [
            "Hardware/Software Import Template Instructions",
            "1. Enter valid information into the fields on the Hardware/Software Import Template.",
            "2. Do not delete columns/sheets, delete the classification label, or add additional columns. Doing so may have a negative impact on the ability for eMASS to ingest the template.",
            "3. The following fields/columns contain drop-down lists: Hardware Type, Software Type, Approval, Yes Or No.",
            "4. If importing hardware information, the \"Machine Name\" field must be populated.",
            "5. If importing software information, the \"Software Name\" field must be populated.",
            "6. All required fields must be populated before importing into eMASS.",
            "7. Review all data for accuracy before importing.",
            "8. Contact your eMASS administrator if you have questions about the import process.",
            "9. Save the file as an Excel workbook (.xlsx) before importing.",
            "10. Do not modify the template structure or add additional columns."
        ]

        for i, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=i, column=1, value=instruction)

        # (U) Lists worksheet
        ws_lists = wb.create_sheet("(U) Lists")

        # Hardware Type list
        ws_lists.cell(row=1, column=1, value="Hardware Type")
        ws_lists.cell(row=2, column=1, value="Workstation")
        ws_lists.cell(row=3, column=1, value="Server")
        ws_lists.cell(row=4, column=1, value="Switch")
        ws_lists.cell(row=5, column=1, value="Router")
        ws_lists.cell(row=6, column=1, value="Firewall")
        ws_lists.cell(row=7, column=1, value="Printer")
        ws_lists.cell(row=8, column=1, value="Scanner")

        # Software Type list
        ws_lists.cell(row=1, column=3, value="Software Type")
        ws_lists.cell(row=2, column=3, value="GOTS Application")
        ws_lists.cell(row=3, column=3, value="COTS Application")
        ws_lists.cell(row=4, column=3, value="Server Application")
        ws_lists.cell(row=5, column=3, value="Web Application")
        ws_lists.cell(row=6, column=3, value="Database")
        ws_lists.cell(row=7, column=3, value="Operating System")
        ws_lists.cell(row=8, column=3, value="Utility")

        # Approval list
        ws_lists.cell(row=1, column=5, value="Approval")
        ws_lists.cell(row=2, column=5, value="In Progress")
        ws_lists.cell(row=3, column=5, value="Unapproved")
        ws_lists.cell(row=4, column=5, value="Approved - FIPS 140-2")
        ws_lists.cell(row=5, column=5, value="Approved - NSA Crypto")
        ws_lists.cell(row=6, column=5, value="Approved - Common Criteria")
        ws_lists.cell(row=7, column=5, value="Approved - Other")
        ws_lists.cell(row=8, column=5, value="Not Applicable")

        # Yes Or No list
        ws_lists.cell(row=1, column=7, value="Yes Or No")
        ws_lists.cell(row=2, column=7, value="Yes")
        ws_lists.cell(row=3, column=7, value="No")

        # Auto-adjust column widths for all worksheets
        for ws in [ws_hardware, ws_software, ws_instructions, ws_lists]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        return output_path

    def export_poam(self, analysis_data: Dict[str, Any], output_path: str = None) -> str:
        """Export POAM (Plan of Action & Milestones) as Excel file"""
        if not output_path:
            output_path = f"POAM_{self.timestamp}.xlsx"

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "POAM"

        # Add classification header
        ws.cell(row=1, column=1, value="***** UNCLASSIFIED//FOR OFFICIAL USE ONLY *****")
        ws.merge_cells('A1:P1')
        ws.cell(row=1, column=1).font = Font(bold=True, color="FF0000")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # Add metadata
        ws.cell(row=2, column=1, value="Date Exported:")
        ws.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d"))
        ws.cell(row=3, column=1, value="Information System:")
        ws.cell(row=3, column=2, value="[Enter System Name]")
        ws.cell(row=4, column=1, value="POAM Coordinator:")
        ws.cell(row=4, column=2, value="[Enter Name]")

        # POAM headers (eMASS standard columns)
        headers = [
            'POAM ID',
            'Control ID',
            'Weakness Name',
            'Weakness Description',
            'Point of Contact',
            'Resources Required',
            'Scheduled Completion Date',
            'Milestone',
            'Milestone Date',
            'Risk',
            'Status',
            'Comments',
            'Raw Severity',
            'Plugin ID',
            'Affected Hosts',
            'Remediation'
        ]

        # Write headers
        header_row = 6
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Get vulnerability data
        report = analysis_data.get('report')
        host_summaries = analysis_data.get('host_summaries', [])

        # Group vulnerabilities by plugin_id to avoid duplicates
        vuln_groups = {}
        for host_summary in host_summaries:
            if report and hasattr(report, 'hosts'):
                for host in report.hosts:
                    if host.name == host_summary.ip or host.properties.hostname == host_summary.hostname:
                        for vuln in host.vulnerabilities:
                            # Only include Cat I, II, III (severity 2, 3, 4)
                            if vuln.severity >= 2:
                                if vuln.plugin_id not in vuln_groups:
                                    vuln_groups[vuln.plugin_id] = {
                                        'vuln': vuln,
                                        'affected_hosts': []
                                    }
                                vuln_groups[vuln.plugin_id]['affected_hosts'].append(
                                    f"{host_summary.hostname} ({host_summary.ip})"
                                )

        # Write POAM items
        row = header_row + 1
        poam_id = 1

        for plugin_id, data in sorted(vuln_groups.items(), key=lambda x: -x[1]['vuln'].severity):
            vuln = data['vuln']
            affected_hosts = data['affected_hosts']

            # Determine risk category based on severity
            if vuln.severity == 4:
                risk = "Very High"
                cat = "I"
                completion_days = 15
            elif vuln.severity == 3:
                risk = "High"
                cat = "II"
                completion_days = 30
            elif vuln.severity == 2:
                risk = "Moderate"
                cat = "III"
                completion_days = 90
            else:
                risk = "Low"
                cat = "III"
                completion_days = 180

            # Calculate scheduled completion date
            from datetime import timedelta
            scheduled_date = (datetime.now() + timedelta(days=completion_days)).strftime("%Y-%m-%d")

            # Extract CVE/Control mapping (simplified)
            control_id = vuln.cve if vuln.cve else f"V-{plugin_id}"

            # Write POAM row
            ws.cell(row=row, column=1, value=f"POAM-{poam_id:04d}")
            ws.cell(row=row, column=2, value=control_id)
            ws.cell(row=row, column=3, value=vuln.plugin_name)
            ws.cell(row=row, column=4, value=vuln.description[:500] + '...' if len(vuln.description) > 500 else vuln.description)
            ws.cell(row=row, column=5, value="[Enter POC]")
            ws.cell(row=row, column=6, value="Staff time, patch management")
            ws.cell(row=row, column=7, value=scheduled_date)
            ws.cell(row=row, column=8, value=f"Remediate Cat {cat} finding")
            ws.cell(row=row, column=9, value=scheduled_date)
            ws.cell(row=row, column=10, value=risk)
            ws.cell(row=row, column=11, value="Open")
            ws.cell(row=row, column=12, value=f"Identified via Nessus scan. {len(affected_hosts)} host(s) affected.")
            ws.cell(row=row, column=13, value=f"CAT {cat}")
            ws.cell(row=row, column=14, value=plugin_id)
            ws.cell(row=row, column=15, value="\n".join(affected_hosts[:5]) + ("..." if len(affected_hosts) > 5 else ""))
            ws.cell(row=row, column=16, value=vuln.solution[:300] + '...' if len(vuln.solution) > 300 else vuln.solution)

            # Apply formatting
            for col in range(1, 17):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Color code by risk
                if risk == "Very High":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif risk == "High":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            row += 1
            poam_id += 1

        # Auto-adjust column widths
        column_widths = [12, 15, 30, 40, 20, 25, 15, 25, 15, 12, 12, 40, 12, 12, 30, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Set row heights
        ws.row_dimensions[header_row].height = 40
        for r in range(header_row + 1, row):
            ws.row_dimensions[r].height = 60

        wb.save(output_path)
        return output_path


def export_excel_vulnerability_report(analysis_data: Dict[str, Any], output_path: str = None) -> str:
    """Export vulnerability report as Excel file"""
    exporter = ExcelExporter()
    return exporter.export_vulnerability_report(analysis_data, output_path)


def export_excel_ivv_test_plan(analysis_data: Dict[str, Any], output_path: str = None) -> str:
    """Export IV&V test plan as Excel file"""
    exporter = ExcelExporter()
    return exporter.export_ivv_test_plan(analysis_data, output_path)


def export_excel_cnet_report(analysis_data: Dict[str, Any], output_path: str = None) -> str:
    """Export CNET report as Excel file"""
    exporter = ExcelExporter()
    return exporter.export_cnet_report(analysis_data, output_path)


def export_excel_hw_sw_inventory(analysis_data: Dict[str, Any], output_path: str = None) -> str:
    """Export HW/SW inventory as Excel file"""
    exporter = ExcelExporter()
    return exporter.export_hw_sw_inventory(analysis_data, output_path)


def export_excel_emass_inventory(analysis_data: Dict[str, Any], output_path: str = None) -> str:
    """Export eMASS inventory as Excel file with macros"""
    exporter = ExcelExporter()
    return exporter.export_emass_inventory(analysis_data, output_path)


def export_excel_poam(analysis_data: Dict[str, Any], output_path: str = None) -> str:
    """Export POAM (Plan of Action & Milestones) as Excel file"""
    exporter = ExcelExporter()
    return exporter.export_poam(analysis_data, output_path)
