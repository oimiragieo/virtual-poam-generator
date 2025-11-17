#!/usr/bin/env python3
"""
Analyze vISSM-generated Excel files to understand their structure
"""

from openpyxl import load_workbook
import os
from datetime import datetime


def analyze_excel_file(filepath):
    """Analyze the structure of an Excel file"""
    print(f"\n{'='*80}")
    print(f"ANALYZING: {os.path.basename(filepath)}")
    print(f"{'='*80}")

    try:
        # Load workbook
        wb = load_workbook(filepath, data_only=True)

        print(f"File size: {os.path.getsize(filepath) / (1024*1024):.1f} MB")
        print(f"Number of worksheets: {len(wb.sheetnames)}")
        print(f"Worksheet names: {wb.sheetnames}")

        # Analyze each worksheet
        for sheet_name in wb.sheetnames:
            print(f"\n--- WORKSHEET: {sheet_name} ---")
            ws = wb[sheet_name]

            # Get dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"Dimensions: {max_row} rows x {max_col} columns")

            # Get first few rows to understand structure
            print("\nFirst 5 rows:")
            for row in range(1, min(6, max_row + 1)):
                row_data = []
                for col in range(1, min(11, max_col + 1)):  # First 10 columns
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value)[:50])  # Truncate long values
                    else:
                        row_data.append("")
                print(f"Row {row}: {row_data}")

            # Look for headers (non-empty cells in first few rows)
            print("\nHeader analysis:")
            for row in range(1, min(4, max_row + 1)):
                headers = []
                for col in range(1, min(21, max_col + 1)):  # First 20 columns
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and str(cell_value).strip():
                        headers.append(str(cell_value).strip())
                if headers:
                    print(f"Row {row} headers: {headers}")

        wb.close()

    except Exception as e:
        print(f"Error analyzing {filepath}: {e}")


def analyze_all_vissm_files():
    """Analyze all vISSM-generated Excel files"""
    files = [
        "vISSM_CET_Report_2021-05-26-2018.xlsx",
        "vISSM_Detailed_Inventory_2021-05-27-0034.xlsx",
        "vISSM_IV&V_Test_Plan_2021-05-25-0720.xlsx",
        "vISSM_Vulnerability_Report_2021-05-26-1959.xlsx",
        "vISSM_eMASS_Inventory_2021-05-25-0126.xlsm",
    ]

    print("vISSM EXCEL FILE ANALYSIS")
    print("=" * 80)
    print(f"Analysis started at: {datetime.now()}")

    for filename in files:
        if os.path.exists(filename):
            analyze_excel_file(filename)
        else:
            print(f"\nFile not found: {filename}")

    print(f"\n{'='*80}")
    print("ANALYSIS COMPLETE")
    print(f"{'='*80}")


if __name__ == "__main__":
    analyze_all_vissm_files()
